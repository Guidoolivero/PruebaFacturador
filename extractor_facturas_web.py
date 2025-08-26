from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import streamlit as st
import pypdf
import pandas as pd
import re
from datetime import datetime
import os
import json
import tempfile
import uuid
import pickle
import pytesseract
from PIL import Image
import io
import numpy as np
import cv2
from pdf2image import convert_from_bytes

# Configuración de variables para los servicios de extracción
USE_PATTERN_MATCHING = True
USE_TESSERACT_OCR = True  # Reemplazo para OpenAI - GRATIS

# Clase para almacenar y gestionar patrones de facturas reconocidos
class PatronesFacturas:
    def __init__(self, ruta_archivo='patrones_facturas.pkl'):
        self.ruta_archivo = ruta_archivo
        self.patrones = self._cargar_patrones()
    
    def _cargar_patrones(self):
        try:
            with open(self.ruta_archivo, 'rb') as f:
                return pickle.load(f)
        except (FileNotFoundError, EOFError):
            return {}
    
    def guardar_patrones(self):
        with open(self.ruta_archivo, 'wb') as f:
            pickle.dump(self.patrones, f)
    
    def encontrar_patron_similar(self, texto, umbral_similitud=0.85):  # Umbral más estricto
        """Encuentra patrones similares basándose en palabras clave, con criterios más estrictos"""
        if not texto or not self.patrones:
            return None
        
        # Extraer palabras clave más distintivas (más largas)
        palabras_clave = set([palabra.lower() for palabra in texto.split() 
                             if len(palabra) > 5])  # Palabras más largas
        
        mejor_patron = None
        mejor_similitud = 0
        
        for id_patron, patron in self.patrones.items():
            palabras_patron = set([palabra.lower() for palabra in patron['texto_muestra'].split() 
                                  if len(palabra) > 5])
            
            if not palabras_patron:
                continue
                
            # Calcular similitud Jaccard
            interseccion = len(palabras_clave.intersection(palabras_patron))
            union = len(palabras_clave.union(palabras_patron))
            
            if union > 0:
                similitud = interseccion / union
                
                # Añadir criterios adicionales
                if similitud > mejor_similitud and similitud >= umbral_similitud:
                    # Verificación adicional: los montos totales deberían ser similares
                    try:
                        total_texto = re.search(r'TOTAL:?\s*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
                        if total_texto and 'estructura' in patron and 'Total' in patron['estructura']:
                            total_patron = float(patron['estructura']['Total'])
                            total_actual = parse_number(total_texto)
                            # Si los totales difieren en más de 10%, no considerar similar
                            if abs(total_actual - total_patron) / max(total_actual, total_patron) > 0.1:
                                continue
                    except:
                        pass  # Si falla la comparación, continuar con la verificación normal
                    
                    mejor_similitud = similitud
                    mejor_patron = patron
        
        return mejor_patron
    
    def agregar_patron(self, datos, texto_muestra, metodo_extraccion):
        id_patron = str(uuid.uuid4())
        self.patrones[id_patron] = {
            'texto_muestra': texto_muestra[:5000],  # Limitar tamaño
            'fecha_creacion': datetime.now().isoformat(),
            'metodo_extraccion': metodo_extraccion,
            'estructura': datos
        }
        self.guardar_patrones()
        return id_patron


# Función para identificar el tipo de factura
def identificar_tipo_factura(texto):
    """
    Identifica el tipo de factura para aplicar patrones específicos.
    Devuelve el tipo de factura como string.
    """
    if ("VIAJES" in texto.upper() or "TURISMO" in texto.upper() or 
        "AGENCIA" in texto.upper() or "PASAJES" in texto.upper() or
        "Srvs de transporte exento" in texto or 
        "FACTURA EMITIDA AL CAMBIO" in texto):
        return "VIAJES"
    
    # Comprobación adicional para FCE_A (Factura Electrónica A)
    if "FCE_A" in texto.upper() or "FCE_B" in texto.upper() or "FACTURA ELECTRONICA" in texto.upper():
        return "ELECTRONICA_AFIP"
    
    if "AFIP" in texto and ("FACTURA ELECTRÓNICA" in texto.upper() or "FACTURA ELECTRONICA" in texto.upper()):
        return "ELECTRONICA_AFIP"
    
    # Facturas B genéricas
    if "FACTURA B" in texto.upper() or "CÓD. 006" in texto or "COD. 006" in texto:
        return "TIPO_B"
    
    # Facturas A genéricas
    if "FACTURA A" in texto.upper() or "CÓD. 001" in texto or "COD. 001" in texto:
        return "TIPO_A"
    
    # Por defecto
    return "GENERICA"

# Función para detectar la moneda de una factura (MEJORADA)
def detectar_moneda(texto):
    """
    Detecta la moneda utilizada en la factura.
    """
    # Prioridad 1: Buscar frases explícitas sobre la moneda
    if re.search(r'emitida\s+en\s+USD', texto, re.IGNORECASE) or \
       re.search(r'pagadera\s+en\s+USD', texto, re.IGNORECASE) or \
       re.search(r'cancelada\s+en\s+(?:USD|dólares|dolares)', texto, re.IGNORECASE) or \
       re.search(r'factura\s+(?:en|de)\s+USD', texto, re.IGNORECASE) or \
       ("cancelada en dicha moneda" in texto and "USD" in texto):
        return "USD"
    
    if re.search(r'emitida\s+en\s+EUR', texto, re.IGNORECASE) or \
       re.search(r'pagadera\s+en\s+EUR', texto, re.IGNORECASE) or \
       re.search(r'cancelada\s+en\s+(?:EUR|euros)', texto, re.IGNORECASE) or \
       re.search(r'factura\s+(?:en|de)\s+EUR', texto, re.IGNORECASE):
        return "EUR"
        
    # Prioridad 2: Buscar símbolos de moneda
    if re.search(r'(?:^|\s)U\$S(?:\s|$)', texto) or \
       re.search(r'(?:^|\s)US\$(?:\s|$)', texto) or \
       re.search(r'\$(?:\s*)USD', texto, re.IGNORECASE) or \
       re.search(r'TOTAL\s+USD', texto, re.IGNORECASE) or \
       re.search(r'USD\s*(?:\d|\.|\,)', texto):
        return "USD"
        
    if re.search(r'€', texto) or \
       re.search(r'EUR(?:\s*)(?:\d|\.|\,)', texto) or \
       re.search(r'TOTAL\s+EUR', texto, re.IGNORECASE):
        return "EUR"
    
    # Prioridad 3: Si hay "TOTAL USD:" con un valor
    if re.search(r'TOTAL\s+USD\s*[:=]?\s*[\d.,]+', texto, re.IGNORECASE):
        return "USD"
        
    if re.search(r'TOTAL\s+EUR\s*[:=]?\s*[\d.,]+', texto, re.IGNORECASE):
        return "EUR"
    
    # Prioridad 4: Detección básica de moneda por defecto
    if "$" in texto and not re.search(r'US\$|U\$S|\$\s*US', texto, re.IGNORECASE):
        # Si hay $ pero no parece ser USD, probablemente sea ARS
        return "ARS"
    
    return "Desconocida"

# Función para convertir texto a número de forma segura
def parse_number(match):
    """Convierte texto de un match regex a número float de manera más robusta."""
    if not match:
        return 0.0
    try:
        # Obtener el texto del match
        value_str = match.group(1).strip()
        
        # Detectar formato de número
        if ',' in value_str and '.' in value_str:
            # Si tiene ambos separadores, determinar cuál es el decimal
            if value_str.rindex('.') > value_str.rindex(','):
                # El punto es el decimal (formato 1,234.56)
                value_str = value_str.replace(',', '')
            else:
                # La coma es el decimal (formato 1.234,56)
                value_str = value_str.replace('.', '').replace(',', '.')
        elif ',' in value_str:
            # Solo tiene comas, verificar si es separador de miles o decimal
            # Si hay 2 o más comas, o la coma está a menos de 3 posiciones del final, es decimal
            if value_str.count(',') >= 2 or len(value_str) - value_str.rindex(',') <= 3:
                value_str = value_str.replace(',', '.')
            else:
                # Es separador de miles
                value_str = value_str.replace(',', '')
        
        # Convertir a float después de procesar
        return float(value_str)
    except (ValueError, AttributeError, IndexError) as e:
        print(f"Error al procesar número '{match.group(1) if match else 'None'}': {str(e)}")
        return 0.0

def detectar_bienes_no_computables(texto, total, gravado, iva, exento):
    """
    Función general para detectar conceptos no gravados en cualquier factura.
    """
    # Lista de frases que indican conceptos no gravados
    frases_no_gravado = [
        "Bienes y srvs. no computables",
        "no computables para la det. del Iva",
        "Conceptos no gravados",
        "No gravado",
        "Operaciones no gravadas",
        "No suj. a IVA",
        "No alcanzado"
    ]
    
    # Verificar si alguna de las frases está presente
    frase_encontrada = None
    for frase in frases_no_gravado:
        if frase.lower() in texto.lower():
            frase_encontrada = frase
            break
    
    # Si encontramos alguna frase indicativa, buscar el valor
    if frase_encontrada:
        # Construir patrones de búsqueda para esta frase específica
        # Patrón 1: La frase seguida de un valor numérico
        patron1 = rf'{re.escape(frase_encontrada)}[^:]*:?\s*[$\s]*(\d[\d.,]+)'
        # Patrón 2: La frase en una tabla seguida de valor
        patron2 = rf'{re.escape(frase_encontrada)}.*?(\d[\d.,]+)'
        # Patrón 3: Búsqueda cerca de la frase
        patron3 = rf'{re.escape(frase_encontrada)}[^$\d]*[$\s]*(\d[\d.,]+)'
        
        # Intentar los patrones en orden
        for patron in [patron1, patron2, patron3]:
            match = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
            if match:
                try:
                    # Limpiar y convertir el valor
                    valor_str = match.group(1).strip()
                    # Manejar diferentes formatos numéricos
                    if ',' in valor_str and '.' in valor_str:
                        if valor_str.rindex('.') > valor_str.rindex(','):
                            valor_str = valor_str.replace(',', '')
                        else:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                    elif ',' in valor_str:
                        valor_str = valor_str.replace(',', '.')
                    
                    valor = float(valor_str)
                    return valor
                except:
                    pass
        
        # Si no se encontró un valor específico pero tenemos la frase y el total
        if total > 0:
            # Calcular suma de componentes conocidos
            suma_componentes = gravado + iva + exento
            
            # Si hay una diferencia significativa entre el total y los componentes conocidos
            if abs(total - suma_componentes) > 1.0 and (total - suma_componentes) > 0:
                return total - suma_componentes
    
    # Si no se encontró ninguna frase específica pero hay diferencia significativa
    if total > 0 and (gravado > 0 or iva > 0 or exento > 0):
        suma_componentes = gravado + iva + exento
        
        # Si hay una diferencia mayor al 15% del total
        if (total - suma_componentes) / total > 0.15:
            return total - suma_componentes
    
    return 0.0
    """
    Función general para detectar conceptos no gravados en cualquier factura.
    """
    # Lista de frases que indican conceptos no gravados
    frases_no_gravado = [
        "Bienes y srvs. no computables",
        "no computables para la det. del Iva",
        "Conceptos no gravados",
        "No gravado",
        "Operaciones no gravadas",
        "No suj. a IVA",
        "No alcanzado"
    ]
    
    # Verificar si alguna de las frases está presente
    frase_encontrada = None
    for frase in frases_no_gravado:
        if frase.lower() in texto.lower():
            frase_encontrada = frase
            break
    
    # Si encontramos alguna frase indicativa, buscar el valor
    if frase_encontrada:
        # Construir patrones de búsqueda para esta frase específica
        # Patrón 1: La frase seguida de un valor numérico
        patron1 = rf'{re.escape(frase_encontrada)}[^:]*:?\s*[$\s]*(\d[\d.,]+)'
        # Patrón 2: La frase en una tabla seguida de valor
        patron2 = rf'{re.escape(frase_encontrada)}.*?(\d[\d.,]+)'
        # Patrón 3: Búsqueda cerca de la frase
        patron3 = rf'{re.escape(frase_encontrada)}[^$\d]*[$\s]*(\d[\d.,]+)'
        
        # Intentar los patrones en orden
        for patron in [patron1, patron2, patron3]:
            match = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
            if match:
                try:
                    # Limpiar y convertir el valor
                    valor_str = match.group(1).strip()
                    # Manejar diferentes formatos numéricos
                    if ',' in valor_str and '.' in valor_str:
                        if valor_str.rindex('.') > valor_str.rindex(','):
                            valor_str = valor_str.replace(',', '')
                        else:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                    elif ',' in valor_str:
                        valor_str = valor_str.replace(',', '.')
                    
                    valor = float(valor_str)
                    return valor
                except:
                    pass
        
        # Si no se encontró un valor específico pero tenemos la frase y el total
        if total > 0:
            # Calcular suma de componentes conocidos
            suma_componentes = gravado + iva + exento
            
            # Si hay una diferencia significativa entre el total y los componentes conocidos
            if abs(total - suma_componentes) > 1.0 and (total - suma_componentes) > 0:
                return total - suma_componentes
    
    # Si no se encontró ninguna frase específica pero hay diferencia significativa
    if total > 0 and (gravado > 0 or iva > 0 or exento > 0):
        suma_componentes = gravado + iva + exento
        
        # Si hay una diferencia mayor al 15% del total
        if (total - suma_componentes) / total > 0.15:
            return total - suma_componentes
    
    return 0.0
    """
    Busca específicamente "Bienes y srvs. no computables" y extrae su valor.
    Si no lo encuentra directamente, intenta calcularlo por diferencia.
    """
    # Búsqueda con patrones más flexibles
    patrones = [
        r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*[\$\s]*([\d.,]+)',
        r'Bienes\s+y\s+srvs\.\s+no\s+computables[^0-9]*(\d[\d.,]+)',
        r'Srvs\s+(?:de\s+)?transporte\s+exento\s+s\/ley[^:]*:?\s*[\$\s]*([\d.,]+)',
        r'no\s+computables\s+para\s+la\s+det\.\s+del\s+Iva:?\s*[\$\s]*([\d.,]+)'
    ]
    
    for patron in patrones:
        match = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
        if match:
            try:
                valor_str = match.group(1).replace('.', '').replace(',', '.')
                return float(valor_str)
            except:
                pass
    
    # Si no encuentra directamente, intentar extraer de tabla
    tabla_match = re.search(r'Bienes y srvs. no computables.*?(\d[\d.,]+)', texto, re.IGNORECASE | re.DOTALL)
    if tabla_match:
        try:
            valor_str = tabla_match.group(1).replace('.', '').replace(',', '.')
            return float(valor_str)
        except:
            pass
    
    # Si tiene la frase pero no encontró el valor, calcular por diferencia
    if "Bienes y srvs. no computables" in texto and total > 0:
        suma_componentes = gravado + iva + exento
        if abs(total - suma_componentes) > 1.0:  # Si hay diferencia significativa
            return total - suma_componentes
    
    # Última verificación: si hay mucha diferencia entre total y componentes
    if total > 0 and (gravado > 0 or iva > 0 or exento > 0):
        suma_componentes = gravado + iva + exento
        # Si hay diferencia mayor al 15% del total
        if (total - suma_componentes) / total > 0.15:
            return total - suma_componentes
    
    return 0.0
# Función específica para extraer datos de facturas de agencias de viajes
def extraer_datos_factura_viajes(texto):
    """
    Extrae datos específicos de facturas de agencias de viajes/turismo.
    """
    # Caso específico para Grupo On Line
    if "GRUPO ON LINE" in texto.upper():
        # Buscar específicamente el valor en la parte inferior de la factura
        total_usd_match = re.search(r'TOTAL\s+USD:\s*([\d.,]+)', texto, re.IGNORECASE)
        
        # Buscar valor de servicios de transporte (aún más específico)
        transporte_match = re.search(r'Srvs\s+de\s+transporte\s+exento\s+s/ley\s+23871:\s*([\d.,]+)', texto)
        
        # Extraer números de factura y fecha
        nro_factura_match = re.search(r'Nro:\s*(\d{4}\s*-\s*\d{5,})', texto)
        fecha_match = re.search(r'Fecha de Emisión:\s*(\d{2}/\d{2}/\d{4})', texto)
        
        # Valor de total directamente de la parte inferior
        total_valor = parse_number(total_usd_match) if total_usd_match else 0.0
        
        # Valor de transporte exento
        transporte_valor = 0.0
        if transporte_match:
            # Tratamiento especial: extraer el número directamente sin usar parse_number
            transporte_str = transporte_match.group(1).strip()
            # Imprimir para depuración
            st.write(f"Valor extraído transporte (raw): '{transporte_str}'")
            try:
                # Manejar específicamente el formato esperado #,###.## 
                if ',' in transporte_str and '.' in transporte_str:
                    # Formato americano con coma como separador de miles
                    transporte_valor = float(transporte_str.replace(',', ''))
                else:
                    transporte_valor = float(transporte_str.replace(',', '.'))
                st.write(f"Transporte convertido: {transporte_valor}")
            except Exception as e:
                st.warning(f"Error al procesar valor de transporte: {str(e)}")
        
        # Si no tenemos total pero sí transporte, usar ese valor
        if total_valor == 0.0 and transporte_valor > 0:
            total_valor = transporte_valor
        
        # Crear resultado con valores encontrados
        return {
            'Numero_Factura': nro_factura_match.group(1) if nro_factura_match else None,
            'Fecha': fecha_match.group(1) if fecha_match else None,
            'No_Gravado': 0.0,
            'Exento': transporte_valor,  # El transporte exento es el valor clave
            'Gravado': 0.0,
            'IVA': 0.0,
            'Total': total_valor,
            'Moneda': 'USD'
        }
    # Detectar número de factura
    nro_factura_patterns = [
        r'Nro:?\s*(\d{4}\s*-\s*\d{5,8})',
        r'Factura\s+(?:[Nn]ro|[Nn][°o])\.?:\s*([A-Z0-9\-]+)',
        r'(?:Comprobante|Factura)\s+[Nn](?:ro|[°o])\.?:\s*([A-Z0-9\-]+)',
    ]
    
    nro_factura = None
    for pattern in nro_factura_patterns:
        match = re.search(pattern, texto, re.IGNORECASE)
        if match:
            nro_factura = match.group(1).strip()
            break
    
    # Buscar fecha en varios formatos
    fecha_patterns = [
        r'Fecha de Emisión:?\s*(\d{1,2}/\d{1,2}/\d{4})',
        r'Fecha:?\s*(\d{1,2}/\d{1,2}/\d{4})',
        r'Emitido el:?\s*(\d{1,2}/\d{1,2}/\d{4})'
    ]
    
    fecha = None
    for pattern in fecha_patterns:
        match = re.search(pattern, texto, re.IGNORECASE)
        if match:
            fecha = match.group(1).strip()
            break
    
    # Buscar el total USD directamente
    total_patterns = [
        r'TOTAL\s+USD:\s*([\d.,]+)',
        r'TOTAL\s+FACTURA\s+USD\s*:\s*([\d.,]+)',
        r'TOTAL(?:\s+GENERAL)?:\s*(?:USD)?\s*([\d.,]+)'
    ]
    
    total = 0.0
    for pattern in total_patterns:
        match = re.search(pattern, texto, re.IGNORECASE)
        if match:
            total = parse_number(match)
            break
    
    # Buscar servicios de transporte exento y otros componentes
    transporte_exento_match = re.search(r'Srvs de transporte exento s/ley 23871:\s*([\d.,]+)', texto)
    transporte_exento = parse_number(transporte_exento_match) if transporte_exento_match else 0.0
    
    # Buscar gravado 21% y 10.5%
    gravado_21_match = re.search(r'Gravado 21%:\s*([\d.,]+)', texto)
    gravado_21 = parse_number(gravado_21_match) if gravado_21_match else 0.0
    
    gravado_10_5_match = re.search(r'Gravado 10\.5%:\s*([\d.,]+)', texto)
    gravado_10_5 = parse_number(gravado_10_5_match) if gravado_10_5_match else 0.0
    
    # Buscar IVA 21% y 10.5%
    iva_21_match = re.search(r'Iva 21%:\s*([\d.,]+)', texto)
    iva_21 = parse_number(iva_21_match) if iva_21_match else 0.0
    
    iva_10_5_match = re.search(r'Iva 10\.5%:\s*([\d.,]+)', texto)
    iva_10_5 = parse_number(iva_10_5_match) if iva_10_5_match else 0.0
    
    # Si no encontramos total pero tenemos otros valores
    if total == 0:
        # Intentar extraer el total de una línea que dice TOTAL USD
        lines = texto.split('\n')
        for line in lines:
            if "TOTAL USD" in line.upper():
                numbers = re.findall(r'[\d.,]+', line)
                if numbers:
                    # El último número en la línea probablemente es el total
                    try:
                        total = float(numbers[-1].replace('.', '').replace(',', '.'))
                    except:
                        pass
    
    # Segunda opción: calcular el total a partir de sus componentes
    if total == 0:
        # Si tenemos servicios de transporte exento, probablemente ese sea el valor principal
        if transporte_exento > 0:
            # Total = transporte exento + gravados + IVAs
            total = transporte_exento + gravado_21 + gravado_10_5 + iva_21 + iva_10_5
    
    # Calcular componentes de la factura (separando los que no están gravados)
    gravado = gravado_21 + gravado_10_5
    iva = iva_21 + iva_10_5
    no_gravado = 0.0
    exento = transporte_exento
    
    # Si todavía no tenemos un total pero tenemos el exento, usarlo como total
    if total == 0 and exento > 0:
        total = exento
    
    # Detectar moneda
    moneda = 'USD' if re.search(r'USD|Moneda:\s*USD', texto, re.IGNORECASE) else 'Desconocida'
    
    return {
        'Numero_Factura': nro_factura,
        'Fecha': fecha,
        'No_Gravado': no_gravado,
        'Exento': exento,
        'Gravado': gravado,
        'IVA': iva,
        'Total': total,
        'Moneda': moneda
    }


def extraer_con_tesseract_ocr(archivo_pdf):
    """
    Extrae datos de facturas usando Tesseract OCR (ALTERNATIVA GRATUITA a OpenAI).
    """
    try:
        # Convertir PDF a imágenes
        with tempfile.NamedTemporaryFile(suffix='.pdf') as temp_pdf:
            temp_pdf.write(archivo_pdf.read())
            temp_pdf.flush()
            
            # Reiniciar el puntero del archivo para usos futuros
            archivo_pdf.seek(0)
            
            # Convertir PDF a imágenes
            images = convert_from_bytes(archivo_pdf.read())
        
        # Procesar cada página con OCR
        texto_completo = ""
        for img in images:
            # Preprocesar la imagen para mejorar el OCR
            img_np = np.array(img)
            img_gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
            img_thresh = cv2.threshold(img_gray, 150, 255, cv2.THRESH_BINARY)[1]
            
            # Aplicar OCR
            texto_pagina = pytesseract.image_to_string(img_thresh, lang='spa')
            texto_completo += texto_pagina + "\n\n"
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto_completo)
        
        # Si es factura de viajes/turismo, usar extracción especializada
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto_completo)
            datos_viajes['Nombre_Archivo'] = archivo_pdf.name
            datos_viajes['Metodo'] = 'Tesseract-OCR-Viajes'
            return datos_viajes
        
        # Para otros tipos de facturas, continuar con el proceso normal
        # Detectar moneda
        moneda = detectar_moneda(texto_completo)
        
        # MODIFICACIÓN: Forzar moneda ARS para facturas argentinas
        if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"] and moneda == 'Desconocida':
            moneda = 'ARS'
        
        # Función para convertir texto a número de forma segura (ya definida anteriormente)
        def parse_number(match):
            if not match:
                return 0.0
            try:
                # Eliminar puntos de miles y reemplazar coma decimal por punto
                value_str = match.group(1).replace('.', '').replace(',', '.')
                return float(value_str)
            except (ValueError, AttributeError):
                return 0.0
        
        # Detectar tipo de factura para usar patrones específicos
        is_factura_b = "CÓD. 006" in texto_completo or "FACTURA B" in texto_completo
        is_factura_a = "CÓD. 001" in texto_completo or "FACTURA A" in texto_completo
        
        # Patrones de expresión regular para diferentes campos (optimizados para OCR)
        # Número de factura
        nro_factura_patterns = [
            r'Comp\.\s*Nro:?\s*(\d+)',
            r'(?:Factura|FACTURA)\s+[Nn](?:[°o]|ro):?\s*([A-Z0-9\-]+)',
            r'[Nn](?:[°o]|ro)\.?\s*(?:Factura|Comprobante):?\s*([A-Z0-9\-]+)',
            r'(?:Factura|Comprobante)\s+(?:[Nn](?:[°o]|ro))?[:\s]+([A-Z0-9\-]+)'
        ]
        
        # Fecha
        fecha_patterns = [
            r'Fecha de Emisi[óo]n:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'Fecha:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'Emitido\s+(?:el)?:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})'
        ]
        
        # NUEVOS patrones para No Gravado
        no_gravado_patterns = [
            r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*\$?\s*([\d.,]+)',
            r'Conceptos\s+no\s+gravados:?\s*\$?\s*([\d.,]+)',
            r'No\s+gravado:?\s*\$?\s*([\d.,]+)',
            r'Importe(?:s)?\s+no\s+gravado(?:s)?:?\s*\$?\s*([\d.,]+)',
            r'Op\.\s+No\s+Gravadas:?\s*\$?\s*([\d.,]+)',
            r'No\s+suj\.\s+a\s+IVA:?\s*\$?\s*([\d.,]+)',
            r'No\s+alcanzado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Total
        total_patterns = [
            r'Importe Total:.*?(\d[\d.,]+)',
            r'TOTAL:?\s*\$?\s*([\d.,]+)',
            r'Total:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:TOTAL|Total):?\s*\$?\s*([\d.,]+)',
            r'(?<!\w)Total(?!\w).*?(\d[\d.,]+)'
        ]
        
        # Exento
        exento_patterns = [
            r'Importe Exento:.*?(\d[\d.,]+)',
            r'Exento:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:EXENTO|Exento):?\s*\$?\s*([\d.,]+)',
            r'Op\.\s+Exentas:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Gravado
        gravado_patterns = [
            r'Importe Neto Gravado:.*?(\d[\d.,]+)',
            r'Gravado:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:NETO\s+)?(?:GRAVADO|Gravado):?\s*\$?\s*([\d.,]+)',
            r'Neto\s+Gravado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # IVA
        iva_patterns = [
            r'IVA\s+21%:.*?(\d[\d.,]+)',
            r'IVA\s+\(?21%\)?:?\s*\$?\s*([\d.,]+)',
            r'IVA:?\s*\$?\s*([\d.,]+)',
            r'I\.V\.A\.(?:\s+\d+%)?:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Aplicar todos los patrones y tomar el primer match para cada campo
        def apply_patterns(patterns, text):
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    return match
            return None
        
        # Buscar primero en el texto completo
        nro_factura_match = apply_patterns(nro_factura_patterns, texto_completo)
        fecha_match = apply_patterns(fecha_patterns, texto_completo)
        no_gravado_match = apply_patterns(no_gravado_patterns, texto_completo)  # NUEVO: Buscar No Gravado
        exento_match = apply_patterns(exento_patterns, texto_completo)
        gravado_match = apply_patterns(gravado_patterns, texto_completo)
        iva_match = apply_patterns(iva_patterns, texto_completo)
        total_match = apply_patterns(total_patterns, texto_completo)
        
        # Extraer valores numéricos
        no_gravado = parse_number(no_gravado_match)  # NUEVO: Extraer valor de No Gravado
        exento = parse_number(exento_match)
        gravado = parse_number(gravado_match)
        iva = parse_number(iva_match)
        total_extraido = parse_number(total_match)
        
        # Si no se encontró el valor de No Gravado específicamente para "Bienes y srvs. no computables"
        if no_gravado == 0 and "Bienes y srvs. no computables" in texto_completo:
            # Buscar específicamente para "Bienes y srvs. no computables" con un patrón más flexible
            bienes_no_comp_match = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*[\$\s]*([\d.,]+)', texto_completo, re.IGNORECASE)
            if bienes_no_comp_match:
                no_gravado = parse_number(bienes_no_comp_match)
            # Si aún no encuentra, buscar cerca de la frase
            else:
                # Buscar números cerca de la frase "Bienes y srvs. no computables"
                contexto = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^$]*(\d[\d.,]+)', texto_completo, re.IGNORECASE)
                if contexto:
                    try:
                        valor_str = re.search(r'(\d[\d.,]+)', contexto.group(1))
                        if valor_str:
                            no_gravado = float(valor_str.group(1).replace('.', '').replace(',', '.'))
                    except:
                        pass
        
        # Si no se encontró el valor exento, intentar buscarlo en tablas
        if exento == 0:
            exento_tabla_match = re.search(r'(?:Exento|EXENTO)\s+(\d[\d.,]+)', texto_completo)
            if exento_tabla_match:
                exento = parse_number(exento_tabla_match)
        
        # Si no se encontró el valor gravado, intentar buscarlo en tablas
        if gravado == 0:
            gravado_tabla_match = re.search(r'(?:Gravado|GRAVADO)\s+(\d[\d.,]+)', texto_completo)
            if gravado_tabla_match:
                gravado = parse_number(gravado_tabla_match)
        
        # Calcular total si no se pudo extraer
        total_calculado = exento + gravado + iva + no_gravado  # Incluir no_gravado en el cálculo
        
        # Comparar total extraído vs calculado y decidir cuál usar
        if total_extraido > 0:
            diferencia = abs(total_extraido - total_calculado)
            if diferencia / (total_extraido + 0.001) > 0.05:  # Diferencia mayor al 5%
                # Si hay gran diferencia, usar el extraído pero alertar
                total = total_extraido
                # Si tenemos total pero los componentes no suman, verificar si hay "Bienes y srvs"
                if "Bienes y srvs. no computables" in texto_completo and no_gravado == 0:
                    # Asignar la diferencia a No Gravado
                    no_gravado = total_extraido - (exento + gravado + iva)
                    no_gravado = max(0, no_gravado)  # Asegurar que no sea negativo
                st.warning(f"Diferencia significativa entre total extraído ({total_extraido}) y calculado ({total_calculado})")
            else:
                total = total_extraido
        else:
            total = total_calculado
        
        # Si tenemos un total pero los componentes suman cero, buscar componentes específicos
        if total > 0 and total_calculado == 0:
            # Si hay "Bienes y srvs. no computables", asignar todo el total a No Gravado
            if "Bienes y srvs. no computables" in texto_completo:
                no_gravado = total
            # Si hay "exento", asignar todo a exento
            elif "exento" in texto_completo.lower():
                exento = total
            # Si hay mención de IVA específico, calcular valores
            elif "21%" in texto_completo:
                gravado = total / 1.21
                iva = gravado * 0.21
            elif "10.5%" in texto_completo:
                gravado = total / 1.105
                iva = gravado * 0.105
            else:
                # Si no hay indicación específica, asignar como No Gravado por defecto
                no_gravado = total
        
        # Asegurar que todos los valores sean números positivos
        no_gravado = max(0, no_gravado)
        exento = max(0, exento)
        gravado = max(0, gravado)
        iva = max(0, iva)
        total = max(0, total)
        
        # Verificación final para Bienes y srvs. no computables
        if no_gravado == 0 or (total > 0 and abs(total - (no_gravado + exento + gravado + iva)) > 1.0):
            no_gravado = detectar_bienes_no_computables(texto_completo, total, gravado, iva, exento)
        
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': nro_factura_match.group(1) if nro_factura_match else None,
            'Fecha': fecha_match.group(1) if fecha_match else None,
            'No_Gravado': no_gravado,  # Ahora incluye el valor capturado
            'Exento': exento,
            'Gravado': gravado,
            'IVA': iva,
            'Total': total,
            'Moneda': moneda,
            'Metodo': 'Tesseract-OCR'
        }
        
    except Exception as e:
        st.warning(f"Error OCR: {str(e)}")
        return None
    """
    Extrae datos de facturas usando Tesseract OCR (ALTERNATIVA GRATUITA a OpenAI).
    """
    try:
        # Convertir PDF a imágenes
        with tempfile.NamedTemporaryFile(suffix='.pdf') as temp_pdf:
            temp_pdf.write(archivo_pdf.read())
            temp_pdf.flush()
            
            # Reiniciar el puntero del archivo para usos futuros
            archivo_pdf.seek(0)
            
            # Convertir PDF a imágenes
            images = convert_from_bytes(archivo_pdf.read())
        
        # Procesar cada página con OCR
        texto_completo = ""
        for img in images:
            # Preprocesar la imagen para mejorar el OCR
            img_np = np.array(img)
            img_gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
            img_thresh = cv2.threshold(img_gray, 150, 255, cv2.THRESH_BINARY)[1]
            
            # Aplicar OCR
            texto_pagina = pytesseract.image_to_string(img_thresh, lang='spa')
            texto_completo += texto_pagina + "\n\n"
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto_completo)
        
        # Si es factura de viajes/turismo, usar extracción especializada
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto_completo)
            datos_viajes['Nombre_Archivo'] = archivo_pdf.name
            datos_viajes['Metodo'] = 'Tesseract-OCR-Viajes'
            return datos_viajes
        
        # Para otros tipos de facturas, continuar con el proceso normal
        # Detectar moneda
        moneda = detectar_moneda(texto_completo)
        
        # MODIFICACIÓN: Forzar moneda ARS para facturas argentinas
        if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"] and moneda == 'Desconocida':
            moneda = 'ARS'
        
        # Función para convertir texto a número de forma segura (ya definida anteriormente)
        def parse_number(match):
            if not match:
                return 0.0
            try:
                # Eliminar puntos de miles y reemplazar coma decimal por punto
                value_str = match.group(1).replace('.', '').replace(',', '.')
                return float(value_str)
            except (ValueError, AttributeError):
                return 0.0
        
        # Detectar tipo de factura para usar patrones específicos
        is_factura_b = "CÓD. 006" in texto_completo or "FACTURA B" in texto_completo
        is_factura_a = "CÓD. 001" in texto_completo or "FACTURA A" in texto_completo
        
        # Patrones de expresión regular para diferentes campos (optimizados para OCR)
        # Número de factura
        nro_factura_patterns = [
            r'Comp\.\s*Nro:?\s*(\d+)',
            r'(?:Factura|FACTURA)\s+[Nn](?:[°o]|ro):?\s*([A-Z0-9\-]+)',
            r'[Nn](?:[°o]|ro)\.?\s*(?:Factura|Comprobante):?\s*([A-Z0-9\-]+)',
            r'(?:Factura|Comprobante)\s+(?:[Nn](?:[°o]|ro))?[:\s]+([A-Z0-9\-]+)'
        ]
        
        # Fecha
        fecha_patterns = [
            r'Fecha de Emisi[óo]n:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'Fecha:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'Emitido\s+(?:el)?:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})'
        ]
        
        # NUEVOS patrones para No Gravado
        no_gravado_patterns = [
            r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*\$?\s*([\d.,]+)',
            r'Conceptos\s+no\s+gravados:?\s*\$?\s*([\d.,]+)',
            r'No\s+gravado:?\s*\$?\s*([\d.,]+)',
            r'Importe(?:s)?\s+no\s+gravado(?:s)?:?\s*\$?\s*([\d.,]+)',
            r'Op\.\s+No\s+Gravadas:?\s*\$?\s*([\d.,]+)',
            r'No\s+suj\.\s+a\s+IVA:?\s*\$?\s*([\d.,]+)',
            r'No\s+alcanzado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Total
        total_patterns = [
            r'Importe Total:.*?(\d[\d.,]+)',
            r'TOTAL:?\s*\$?\s*([\d.,]+)',
            r'Total:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:TOTAL|Total):?\s*\$?\s*([\d.,]+)',
            r'(?<!\w)Total(?!\w).*?(\d[\d.,]+)'
        ]
        
        # Exento
        exento_patterns = [
            r'Importe Exento:.*?(\d[\d.,]+)',
            r'Exento:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:EXENTO|Exento):?\s*\$?\s*([\d.,]+)',
            r'Op\.\s+Exentas:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Gravado
        gravado_patterns = [
            r'Importe Neto Gravado:.*?(\d[\d.,]+)',
            r'Gravado:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:NETO\s+)?(?:GRAVADO|Gravado):?\s*\$?\s*([\d.,]+)',
            r'Neto\s+Gravado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # IVA
        iva_patterns = [
            r'IVA\s+21%:.*?(\d[\d.,]+)',
            r'IVA\s+\(?21%\)?:?\s*\$?\s*([\d.,]+)',
            r'IVA:?\s*\$?\s*([\d.,]+)',
            r'I\.V\.A\.(?:\s+\d+%)?:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Aplicar todos los patrones y tomar el primer match para cada campo
        def apply_patterns(patterns, text):
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    return match
            return None
        
        # Buscar primero en el texto completo
        nro_factura_match = apply_patterns(nro_factura_patterns, texto_completo)
        fecha_match = apply_patterns(fecha_patterns, texto_completo)
        no_gravado_match = apply_patterns(no_gravado_patterns, texto_completo)  # NUEVO: Buscar No Gravado
        exento_match = apply_patterns(exento_patterns, texto_completo)
        gravado_match = apply_patterns(gravado_patterns, texto_completo)
        iva_match = apply_patterns(iva_patterns, texto_completo)
        total_match = apply_patterns(total_patterns, texto_completo)
        
        # Extraer valores numéricos
        no_gravado = parse_number(no_gravado_match)  # NUEVO: Extraer valor de No Gravado
        exento = parse_number(exento_match)
        gravado = parse_number(gravado_match)
        iva = parse_number(iva_match)
        total_extraido = parse_number(total_match)
        
        # Si no se encontró el valor de No Gravado específicamente para "Bienes y srvs. no computables"
        if no_gravado == 0 and "Bienes y srvs. no computables" in texto_completo:
            # Buscar específicamente para "Bienes y srvs. no computables" con un patrón más flexible
            bienes_no_comp_match = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*[\$\s]*([\d.,]+)', texto_completo, re.IGNORECASE)
            if bienes_no_comp_match:
                no_gravado = parse_number(bienes_no_comp_match)
            # Si aún no encuentra, buscar cerca de la frase
            else:
                # Buscar números cerca de la frase "Bienes y srvs. no computables"
                contexto = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^$]*(\d[\d.,]+)', texto_completo, re.IGNORECASE)
                if contexto:
                    try:
                        valor_str = re.search(r'(\d[\d.,]+)', contexto.group(1))
                        if valor_str:
                            no_gravado = float(valor_str.group(1).replace('.', '').replace(',', '.'))
                    except:
                        pass
        
        # Si no se encontró el valor exento, intentar buscarlo en tablas
        if exento == 0:
            exento_tabla_match = re.search(r'(?:Exento|EXENTO)\s+(\d[\d.,]+)', texto_completo)
            if exento_tabla_match:
                exento = parse_number(exento_tabla_match)
        
        # Si no se encontró el valor gravado, intentar buscarlo en tablas
        if gravado == 0:
            gravado_tabla_match = re.search(r'(?:Gravado|GRAVADO)\s+(\d[\d.,]+)', texto_completo)
            if gravado_tabla_match:
                gravado = parse_number(gravado_tabla_match)
        
        # Calcular total si no se pudo extraer
        total_calculado = exento + gravado + iva + no_gravado  # Incluir no_gravado en el cálculo
        
        # Comparar total extraído vs calculado y decidir cuál usar
        if total_extraido > 0:
            diferencia = abs(total_extraido - total_calculado)
            if diferencia / (total_extraido + 0.001) > 0.05:  # Diferencia mayor al 5%
                # Si hay gran diferencia, usar el extraído pero alertar
                total = total_extraido
                # Si tenemos total pero los componentes no suman, verificar si hay "Bienes y srvs"
                if "Bienes y srvs. no computables" in texto_completo and no_gravado == 0:
                    # Asignar la diferencia a No Gravado
                    no_gravado = total_extraido - (exento + gravado + iva)
                    no_gravado = max(0, no_gravado)  # Asegurar que no sea negativo
                st.warning(f"Diferencia significativa entre total extraído ({total_extraido}) y calculado ({total_calculado})")
            else:
                total = total_extraido
        else:
            total = total_calculado
        
        # Si tenemos un total pero los componentes suman cero, buscar componentes específicos
        if total > 0 and total_calculado == 0:
            # Si hay "Bienes y srvs. no computables", asignar todo el total a No Gravado
            if "Bienes y srvs. no computables" in texto_completo:
                no_gravado = total
            # Si hay "exento", asignar todo a exento
            elif "exento" in texto_completo.lower():
                exento = total
            # Si hay mención de IVA específico, calcular valores
            elif "21%" in texto_completo:
                gravado = total / 1.21
                iva = gravado * 0.21
            elif "10.5%" in texto_completo:
                gravado = total / 1.105
                iva = gravado * 0.105
            else:
                # Si no hay indicación específica, asignar como No Gravado por defecto
                no_gravado = total
        
        # Asegurar que todos los valores sean números positivos
        no_gravado = max(0, no_gravado)
        exento = max(0, exento)
        gravado = max(0, gravado)
        iva = max(0, iva)
        total = max(0, total)
        
        # Verificación final para Bienes y srvs. no computables
        if no_gravado == 0 or (total > 0 and abs(total - (no_gravado + exento + gravado + iva)) > 1.0):
            no_gravado = detectar_bienes_no_computables(texto_completo, total, gravado, iva, exento)
        
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': nro_factura_match.group(1) if nro_factura_match else None,
            'Fecha': fecha_match.group(1) if fecha_match else None,
            'No_Gravado': no_gravado,  # Ahora incluye el valor capturado
            'Exento': exento,
            'Gravado': gravado,
            'IVA': iva,
            'Total': total,
            'Moneda': moneda,
            'Metodo': 'Tesseract-OCR'
        }
        
    except Exception as e:
        st.warning(f"Error OCR: {str(e)}")
        return None

def extraer_con_regex(archivo_pdf):
    """
    Extrae datos de facturas usando expresiones regulares avanzadas.
    """
    try:
        # Leer el PDF
        reader = pypdf.PdfReader(archivo_pdf)
        texto = ""
        for pagina in reader.pages:
            texto += pagina.extract_text()
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto)
        
        # Si es factura de viajes/turismo, usar extracción especializada
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto)
            datos_viajes['Nombre_Archivo'] = archivo_pdf.name
            datos_viajes['Metodo'] = 'RegEx-Viajes'
            return datos_viajes
            
        # Detectar moneda
        moneda = detectar_moneda(texto)
        
        # MODIFICACIÓN: Forzar moneda ARS para facturas argentinas
        if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"] and moneda == 'Desconocida':
            moneda = 'ARS'
        
        # Función para convertir texto a número de forma segura
        def parse_number(match):
            if not match:
                return 0.0
            try:
                # Eliminar puntos de miles y reemplazar coma decimal por punto
                value_str = match.group(1).replace('.', '').replace(',', '.')
                return float(value_str)
            except (ValueError, AttributeError):
                return 0.0
        
        # Detectar tipo de factura para usar patrones específicos
        is_factura_b = "CÓD. 006" in texto or "FACTURA B" in texto
        is_factura_a = "CÓD. 001" in texto or "FACTURA A" in texto
        
        # Patrones de expresión regular para diferentes campos
        # Número de factura
        nro_factura_patterns = [
            r'Comp\. Nro:\s*(\d+)',
            r'Factura\s+[Nn][°o]:\s*([A-Z0-9\-]+)',
            r'N[°o]\s*(?:Factura|Comprobante):\s*([A-Z0-9\-]+)',
            r'(?:Factura|Comprobante)\s+(?:[Nn][°o])?[:\s]+([A-Z0-9\-]+)'
        ]
        
        # Fecha
        fecha_patterns = [
            r'Fecha de Emisión:\s*(\d{2}/\d{2}/\d{4})',
            r'Fecha:\s*(\d{2}/\d{2}/\d{4})',
            r'Fecha\s+(?:de\s+)?(?:Emisión|Emision):\s*(\d{2}[-/]\d{2}[-/]\d{4})',
            r'Emitido\s+(?:el)?:\s*(\d{2}[-/]\d{2}[-/]\d{4})'
        ]
        
        # NUEVOS patrones para No Gravado - AGREGADOS para capturar "Bienes y srvs. no computables"
        no_gravado_patterns = [
            r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:\s*\$?\s*([\d.,]+)',
            r'Conceptos\s+no\s+gravados:\s*\$?\s*([\d.,]+)',
            r'No\s+gravado:\s*\$?\s*([\d.,]+)',
            r'Importe(?:s)?\s+no\s+gravado(?:s)?:\s*\$?\s*([\d.,]+)',
            r'Op\.\s+No\s+Gravadas:\s*\$?\s*([\d.,]+)',
            r'No\s+suj\.\s+a\s+IVA:\s*\$?\s*([\d.,]+)',
            r'No\s+alcanzado:\s*\$?\s*([\d.,]+)'
        ]
        
        # Exento
        exento_patterns = [
            r'Importe Exento:.*?(\d[\d.,]+)',
            r'Exento:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:EXENTO|Exento):?\s*\$?\s*([\d.,]+)',
            r'Op.\s+Exentas:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Gravado
        gravado_patterns = [
            r'Importe Neto Gravado:.*?(\d[\d.,]+)',
            r'Gravado:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:NETO\s+)?(?:GRAVADO|Gravado):?\s*\$?\s*([\d.,]+)',
            r'Neto\s+Gravado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # IVA
        iva_patterns = [
            r'IVA 21%:.*?(\d[\d.,]+)',
            r'IVA:?\s*\$?\s*([\d.,]+)',
            r'I\.V\.A\.(?:\s+\d+%)?:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Total
        total_patterns = [
            r'Importe Total:.*?(\d[\d.,]+)',
            r'TOTAL:?\s*\$?\s*([\d.,]+)',
            r'Total:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:TOTAL|Total):?\s*\$?\s*([\d.,]+)',
            r'(?<!\w)Total(?!\w).*?(\d[\d.,]+)'
        ]
        
        # Aplicar todos los patrones y tomar el primer match para cada campo
        def apply_patterns(patterns, text):
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    return match
            return None
        
        # Buscar los patrones en el texto
        nro_factura_match = apply_patterns(nro_factura_patterns, texto)
        fecha_match = apply_patterns(fecha_patterns, texto)
        no_gravado_match = apply_patterns(no_gravado_patterns, texto)  # NUEVO: Buscar No Gravado
        exento_match = apply_patterns(exento_patterns, texto)
        gravado_match = apply_patterns(gravado_patterns, texto)
        iva_match = apply_patterns(iva_patterns, texto)
        total_match = apply_patterns(total_patterns, texto)
        
        # Extraer valores numéricos
        no_gravado = parse_number(no_gravado_match)  # NUEVO: Extraer valor de No Gravado
        exento = parse_number(exento_match)
        gravado = parse_number(gravado_match)
        iva = parse_number(iva_match)
        total_extraido = parse_number(total_match)
        
        # Si no se encontró el valor de No Gravado específicamente para "Bienes y srvs. no computables"
        if no_gravado == 0 and "Bienes y srvs. no computables" in texto:
            # Buscar específicamente para "Bienes y srvs. no computables" con un patrón más flexible
            bienes_no_comp_match = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*[\$\s]*([\d.,]+)', texto, re.IGNORECASE)
            if bienes_no_comp_match:
                no_gravado = parse_number(bienes_no_comp_match)
            # Si aún no encuentra, buscar con un patrón más genérico
            elif "Srvs de transporte exento" in texto:
                srvs_transport_match = re.search(r'Srvs\s+de\s+transporte\s+exento\s+s/ley\s+\d+:\s*([\d.,]+)', texto)
                if srvs_transport_match:
                    no_gravado = parse_number(srvs_transport_match)
        
        # Si no se encontró el valor exento, intentar buscarlo en tablas
        if exento == 0:
            # Buscar en tablas con patrones específicos
            exento_tabla_match = re.search(r'Exento\s+(\d[\d.,]+)', texto)
            if exento_tabla_match:
                exento = parse_number(exento_tabla_match)
        
        # Si no se encontró el valor gravado, intentar buscarlo en tablas
        if gravado == 0:
            # Buscar en tablas con patrones específicos
            gravado_tabla_match = re.search(r'Gravado\s+(\d[\d.,]+)', texto)
            if gravado_tabla_match:
                gravado = parse_number(gravado_tabla_match)
        
        # Calcular total si no se pudo extraer
        total_calculado = no_gravado + exento + gravado + iva  # Actualizado para incluir no_gravado
        
        # Comparar total extraído vs calculado y decidir cuál usar
        if total_extraido > 0:
            diferencia = abs(total_extraido - total_calculado)
            if diferencia / (total_extraido + 0.001) > 0.05:  # Diferencia mayor al 5%
                # Si hay gran diferencia, usar el extraído pero alertar
                total = total_extraido
                # Si tenemos total pero los componentes no suman, verificar si hay "Bienes y srvs"
                if "Bienes y srvs. no computables" in texto and no_gravado == 0:
                    # Asignar la diferencia a No Gravado
                    no_gravado = total_extraido - (exento + gravado + iva)
                    no_gravado = max(0, no_gravado)  # Asegurar que no sea negativo
            else:
                total = total_extraido
        else:
            total = total_calculado
        
        # Si tenemos un total pero los componentes suman cero, buscar componentes específicos
        if total > 0 and total_calculado == 0:
            # Si hay "Bienes y srvs. no computables", asignar todo el total a No Gravado
            if "Bienes y srvs. no computables" in texto:
                no_gravado = total
            # Si hay "exento", asignar todo a exento
            elif "exento" in texto.lower():
                exento = total
            # Si hay mención de IVA específico, calcular valores
            elif "21%" in texto:
                gravado = total / 1.21
                iva = gravado * 0.21
            elif "10.5%" in texto:
                gravado = total / 1.105
                iva = gravado * 0.105
            else:
                # Si no hay indicación específica, asignar como No Gravado por defecto
                no_gravado = total
                
        # Asegurar que todos los valores sean números positivos
        no_gravado = max(0, no_gravado)
        exento = max(0, exento)
        gravado = max(0, gravado)
        iva = max(0, iva)
        total = max(0, total)
        
        # Verificación final para Bienes y srvs. no computables
        if no_gravado == 0 or (total > 0 and abs(total - (no_gravado + exento + gravado + iva)) > 1.0):
            no_gravado = detectar_bienes_no_computables(texto, total, gravado, iva, exento)
        
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': nro_factura_match.group(1) if nro_factura_match else None,
            'Fecha': fecha_match.group(1) if fecha_match else None,
            'No_Gravado': no_gravado,  # Ahora incluye el valor capturado
            'Exento': exento,
            'Gravado': gravado,
            'IVA': iva,
            'Total': total,
            'Moneda': moneda,
            'Metodo': 'RegEx'
        }
        
    except Exception as e:
        st.warning(f"RegEx: {str(e)}")
        return None
    
    """
    Extrae datos de facturas usando expresiones regulares avanzadas.
    """
    try:
        # Leer el PDF
        reader = pypdf.PdfReader(archivo_pdf)
        texto = ""
        for pagina in reader.pages:
            texto += pagina.extract_text()
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto)
        
        # Si es factura de viajes/turismo, usar extracción especializada
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto)
            datos_viajes['Nombre_Archivo'] = archivo_pdf.name
            datos_viajes['Metodo'] = 'RegEx-Viajes'
            return datos_viajes
            
        # Detectar moneda
        moneda = detectar_moneda(texto)
        
        # MODIFICACIÓN: Forzar moneda ARS para facturas argentinas
        if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"] and moneda == 'Desconocida':
            moneda = 'ARS'
        
        # Función para convertir texto a número de forma segura
        def parse_number(match):
            if not match:
                return 0.0
            try:
                # Eliminar puntos de miles y reemplazar coma decimal por punto
                value_str = match.group(1).replace('.', '').replace(',', '.')
                return float(value_str)
            except (ValueError, AttributeError):
                return 0.0
        
        # Detectar tipo de factura para usar patrones específicos
        is_factura_b = "CÓD. 006" in texto or "FACTURA B" in texto
        is_factura_a = "CÓD. 001" in texto or "FACTURA A" in texto
        
        # Patrones de expresión regular para diferentes campos
        # Número de factura
        nro_factura_patterns = [
            r'Comp\. Nro:\s*(\d+)',
            r'Factura\s+[Nn][°o]:\s*([A-Z0-9\-]+)',
            r'N[°o]\s*(?:Factura|Comprobante):\s*([A-Z0-9\-]+)',
            r'(?:Factura|Comprobante)\s+(?:[Nn][°o])?[:\s]+([A-Z0-9\-]+)'
        ]
        
        # Fecha
        fecha_patterns = [
            r'Fecha de Emisión:\s*(\d{2}/\d{2}/\d{4})',
            r'Fecha:\s*(\d{2}/\d{2}/\d{4})',
            r'Fecha\s+(?:de\s+)?(?:Emisión|Emision):\s*(\d{2}[-/]\d{2}[-/]\d{4})',
            r'Emitido\s+(?:el)?:\s*(\d{2}[-/]\d{2}[-/]\d{4})'
        ]
        
        # NUEVOS patrones para No Gravado - AGREGADOS para capturar "Bienes y srvs. no computables"
        no_gravado_patterns = [
            r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:\s*\$?\s*([\d.,]+)',
            r'Conceptos\s+no\s+gravados:\s*\$?\s*([\d.,]+)',
            r'No\s+gravado:\s*\$?\s*([\d.,]+)',
            r'Importe(?:s)?\s+no\s+gravado(?:s)?:\s*\$?\s*([\d.,]+)',
            r'Op\.\s+No\s+Gravadas:\s*\$?\s*([\d.,]+)',
            r'No\s+suj\.\s+a\s+IVA:\s*\$?\s*([\d.,]+)',
            r'No\s+alcanzado:\s*\$?\s*([\d.,]+)'
        ]
        
        # Exento
        exento_patterns = [
            r'Importe Exento:.*?(\d[\d.,]+)',
            r'Exento:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:EXENTO|Exento):?\s*\$?\s*([\d.,]+)',
            r'Op.\s+Exentas:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Gravado
        gravado_patterns = [
            r'Importe Neto Gravado:.*?(\d[\d.,]+)',
            r'Gravado:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:NETO\s+)?(?:GRAVADO|Gravado):?\s*\$?\s*([\d.,]+)',
            r'Neto\s+Gravado:?\s*\$?\s*([\d.,]+)'
        ]
        
        # IVA
        iva_patterns = [
            r'IVA 21%:.*?(\d[\d.,]+)',
            r'IVA:?\s*\$?\s*([\d.,]+)',
            r'I\.V\.A\.(?:\s+\d+%)?:?\s*\$?\s*([\d.,]+)'
        ]
        
        # Total
        total_patterns = [
            r'Importe Total:.*?(\d[\d.,]+)',
            r'TOTAL:?\s*\$?\s*([\d.,]+)',
            r'Total:?\s*\$?\s*([\d.,]+)',
            r'(?:IMPORTE|Importe)\s+(?:TOTAL|Total):?\s*\$?\s*([\d.,]+)',
            r'(?<!\w)Total(?!\w).*?(\d[\d.,]+)'
        ]
        
        # Aplicar todos los patrones y tomar el primer match para cada campo
        def apply_patterns(patterns, text):
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    return match
            return None
        
        # Buscar los patrones en el texto
        nro_factura_match = apply_patterns(nro_factura_patterns, texto)
        fecha_match = apply_patterns(fecha_patterns, texto)
        no_gravado_match = apply_patterns(no_gravado_patterns, texto)  # NUEVO: Buscar No Gravado
        exento_match = apply_patterns(exento_patterns, texto)
        gravado_match = apply_patterns(gravado_patterns, texto)
        iva_match = apply_patterns(iva_patterns, texto)
        total_match = apply_patterns(total_patterns, texto)
        
        # Extraer valores numéricos
        no_gravado = parse_number(no_gravado_match)  # NUEVO: Extraer valor de No Gravado
        exento = parse_number(exento_match)
        gravado = parse_number(gravado_match)
        iva = parse_number(iva_match)
        total_extraido = parse_number(total_match)
        
        # Si no se encontró el valor de No Gravado específicamente para "Bienes y srvs. no computables"
        if no_gravado == 0 and "Bienes y srvs. no computables" in texto:
            # Buscar específicamente para "Bienes y srvs. no computables" con un patrón más flexible
            bienes_no_comp_match = re.search(r'Bienes\s+y\s+srvs\.\s+no\s+computables[^:]*:?\s*[\$\s]*([\d.,]+)', texto, re.IGNORECASE)
            if bienes_no_comp_match:
                no_gravado = parse_number(bienes_no_comp_match)
            # Si aún no encuentra, buscar con un patrón más genérico
            elif "Srvs de transporte exento" in texto:
                srvs_transport_match = re.search(r'Srvs\s+de\s+transporte\s+exento\s+s/ley\s+\d+:\s*([\d.,]+)', texto)
                if srvs_transport_match:
                    no_gravado = parse_number(srvs_transport_match)
        
        # Si no se encontró el valor exento, intentar buscarlo en tablas
        if exento == 0:
            # Buscar en tablas con patrones específicos
            exento_tabla_match = re.search(r'Exento\s+(\d[\d.,]+)', texto)
            if exento_tabla_match:
                exento = parse_number(exento_tabla_match)
        
        # Si no se encontró el valor gravado, intentar buscarlo en tablas
        if gravado == 0:
            # Buscar en tablas con patrones específicos
            gravado_tabla_match = re.search(r'Gravado\s+(\d[\d.,]+)', texto)
            if gravado_tabla_match:
                gravado = parse_number(gravado_tabla_match)
        
        # Calcular total si no se pudo extraer
        total_calculado = no_gravado + exento + gravado + iva  # Actualizado para incluir no_gravado
        
        # Comparar total extraído vs calculado y decidir cuál usar
        if total_extraido > 0:
            diferencia = abs(total_extraido - total_calculado)
            if diferencia / (total_extraido + 0.001) > 0.05:  # Diferencia mayor al 5%
                # Si hay gran diferencia, usar el extraído pero alertar
                total = total_extraido
                # Si tenemos total pero los componentes no suman, verificar si hay "Bienes y srvs"
                if "Bienes y srvs. no computables" in texto and no_gravado == 0:
                    # Asignar la diferencia a No Gravado
                    no_gravado = total_extraido - (exento + gravado + iva)
                    no_gravado = max(0, no_gravado)  # Asegurar que no sea negativo
            else:
                total = total_extraido
        else:
            total = total_calculado
        
        # Si tenemos un total pero los componentes suman cero, buscar componentes específicos
        if total > 0 and total_calculado == 0:
            # Si hay "Bienes y srvs. no computables", asignar todo el total a No Gravado
            if "Bienes y srvs. no computables" in texto:
                no_gravado = total
            # Si hay "exento", asignar todo a exento
            elif "exento" in texto.lower():
                exento = total
            # Si hay mención de IVA específico, calcular valores
            elif "21%" in texto:
                gravado = total / 1.21
                iva = gravado * 0.21
            elif "10.5%" in texto:
                gravado = total / 1.105
                iva = gravado * 0.105
            else:
                # Si no hay indicación específica, asignar como No Gravado por defecto
                no_gravado = total
                
        # Asegurar que todos los valores sean números positivos
        no_gravado = max(0, no_gravado)
        exento = max(0, exento)
        gravado = max(0, gravado)
        iva = max(0, iva)
        total = max(0, total)
        
        # Verificación final para Bienes y srvs. no computables
        if no_gravado == 0 or (total > 0 and abs(total - (no_gravado + exento + gravado + iva)) > 1.0):
            no_gravado = detectar_bienes_no_computables(texto, total, gravado, iva, exento)
        
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': nro_factura_match.group(1) if nro_factura_match else None,
            'Fecha': fecha_match.group(1) if fecha_match else None,
            'No_Gravado': no_gravado,  # Ahora incluye el valor capturado
            'Exento': exento,
            'Gravado': gravado,
            'IVA': iva,
            'Total': total,
            'Moneda': moneda,
            'Metodo': 'RegEx'
        }
        
    except Exception as e:
        st.warning(f"RegEx: {str(e)}")
        return None   

def extraer_datos_pdf(archivo_pdf, patrones_manager, ignore_patterns=False):
    """
    Sistema de cascada para extracción de datos de facturas.
    Intenta varios métodos en orden de preferencia.
    """
    try:
        # Verificar el nombre del archivo para facturas argentinas
        nombre_archivo = archivo_pdf.name.upper()
        es_factura_argentina = False
        es_factura_usd = False
        
        # Verificación rápida del nombre para facturas argentinas
        if "FCE_A" in nombre_archivo or "ROSSO" in nombre_archivo or "FACTURA_A" in nombre_archivo or "AFIP" in nombre_archivo:
            es_factura_argentina = True
            moneda_inicial = "ARS"  # Forzar ARS para este tipo de archivo
        
        # Verificación específica para facturas que sabemos son USD
        if "THAU" in nombre_archivo or "FUENTES" in nombre_archivo:
            es_factura_usd = True
            moneda_inicial = "USD"  # Marcar inicialmente como USD
        
        # Leer el texto para buscar patrones similares
        reader = pypdf.PdfReader(archivo_pdf)
        texto = ""
        for pagina in reader.pages:
            texto += pagina.extract_text()
        
        # Verificación explícita para facturas en USD basada en el texto
        if "emitida en USD" in texto or "TOTAL USD:" in texto or (
            "cancelada en dicha moneda" in texto and "USD" in texto):
            es_factura_usd = True
            moneda_inicial = "USD"
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto)
        
        # Determinar moneda inicial si no se hizo por verificaciones anteriores
        if not es_factura_argentina and not es_factura_usd:
            if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                moneda_inicial = "ARS"  # Forzar ARS para facturas A, B o electrónicas AFIP
            else:
                # Intentar detectar moneda normalmente
                moneda_inicial = detectar_moneda(texto)
        
        # 0. Buscar si hay un patrón similar ya aprendido (solo si no se ignoran patrones)
        if USE_PATTERN_MATCHING and not ignore_patterns:
            patron_similar = patrones_manager.encontrar_patron_similar(texto)
            if patron_similar:
                st.info(f"Usando patrón aprendido previamente")
                estructura = patron_similar['estructura']
                estructura['Nombre_Archivo'] = archivo_pdf.name
                
                # Si el patrón no tiene moneda, intentar asignar una
                if 'Moneda' not in estructura:
                    estructura['Moneda'] = moneda_inicial
                
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    estructura['Moneda'] = 'USD'
                elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                    estructura['Moneda'] = 'ARS'
                    
                # Verificación final para conceptos no gravados
                if estructura['Total'] > 0 and "Bienes y srvs. no computables" in texto:
                    suma_actual = estructura['No_Gravado'] + estructura['Exento'] + estructura['Gravado'] + estructura['IVA']
                    if abs(estructura['Total'] - suma_actual) > 1.0 and estructura['No_Gravado'] < 1.0:
                        estructura['No_Gravado'] = detectar_bienes_no_computables(
                            texto, estructura['Total'], estructura['Gravado'], 
                            estructura['IVA'], estructura['Exento']
                        )
                    
                estructura['Metodo'] = f"Patrón-{patron_similar['metodo_extraccion']}"
                return estructura
        
        # 1. Intentar con Tesseract OCR (alternativa gratuita a OpenAI)
        if USE_TESSERACT_OCR:
            archivo_pdf.seek(0)  # Reiniciar el puntero del archivo
            datos = extraer_con_tesseract_ocr(archivo_pdf)
            if datos and datos['Total'] > 0:
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    datos['Moneda'] = 'USD'
                elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                    datos['Moneda'] = 'ARS'
                else:
                    # Si no hay detección específica, usar la moneda inicial
                    datos['Moneda'] = moneda_inicial
                
                # Guardar el patrón aprendido
                patrones_manager.agregar_patron(datos, texto, 'TesseractOCR')
                return datos
        
        # 2. Intentar con Expresiones Regulares mejoradas
        archivo_pdf.seek(0)  # Reiniciar el puntero del archivo
        datos = extraer_con_regex(archivo_pdf)
        if datos and datos['Total'] > 0:
            # Forzar moneda según detecciones específicas
            if es_factura_usd:
                datos['Moneda'] = 'USD'
            elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                datos['Moneda'] = 'ARS'
            else:
                # Si no hay detección específica, usar la moneda inicial
                datos['Moneda'] = moneda_inicial
            
            # Guardar el patrón aprendido
            patrones_manager.agregar_patron(datos, texto, 'RegEx')
            return datos
        
        # 3. Intentar específicamente con el extractor de facturas de viajes si es ese tipo
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto)
            if datos_viajes and datos_viajes['Total'] > 0:
                datos_viajes['Nombre_Archivo'] = archivo_pdf.name
                
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    datos_viajes['Moneda'] = 'USD'
                elif es_factura_argentina:
                    datos_viajes['Moneda'] = 'ARS'
                else:
                    datos_viajes['Moneda'] = moneda_inicial
                
                datos_viajes['Metodo'] = 'Especializado-Viajes'
                patrones_manager.agregar_patron(datos_viajes, texto, 'Especializado-Viajes')
                return datos_viajes
        
        # 4. Si todo falla, devolver datos vacíos
        resultado_fallido = {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': None,
            'Fecha': None,
            'Total': 0.0,
            'Gravado': 0.0,
            'IVA': 0.0,
            'No_Gravado': 0.0,
            'Exento': 0.0,
            'Moneda': moneda_inicial,  # Al menos tenemos una moneda detectada
            'Metodo': 'Fallido'
        }
        
        # Forzar moneda según detecciones específicas incluso si falló
        if es_factura_usd:
            resultado_fallido['Moneda'] = 'USD'
        elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
            resultado_fallido['Moneda'] = 'ARS'
        
        # Verificación final de consistencia para conceptos no gravados
        if "Bienes y srvs. no computables" in texto:
            valor_no_gravado = detectar_bienes_no_computables(
                texto, resultado_fallido['Total'], resultado_fallido['Gravado'],
                resultado_fallido['IVA'], resultado_fallido['Exento']
            )
            if valor_no_gravado > 0:
                resultado_fallido['No_Gravado'] = valor_no_gravado
            
        return resultado_fallido
        
    except Exception as e:
        st.error(f"Error al procesar {archivo_pdf.name}: {str(e)}")
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': None,
            'Fecha': None,
            'Total': 0.0,
            'Gravado': 0.0,
            'IVA': 0.0,
            'No_Gravado': 0.0,
            'Exento': 0.0,
            'Moneda': 'Desconocida',
            'Metodo': 'Error'
        }
    """
    Sistema de cascada para extracción de datos de facturas.
    Intenta varios métodos en orden de preferencia.
    """
    try:
        # Verificar el nombre del archivo para facturas argentinas
        nombre_archivo = archivo_pdf.name.upper()
        es_factura_argentina = False
        es_factura_usd = False
        
        # Verificación rápida del nombre para facturas argentinas
        if "FCE_A" in nombre_archivo or "ROSSO" in nombre_archivo or "FACTURA_A" in nombre_archivo or "AFIP" in nombre_archivo:
            es_factura_argentina = True
            moneda_inicial = "ARS"  # Forzar ARS para este tipo de archivo
        
        # Verificación específica para facturas que sabemos son USD
        if "THAU" in nombre_archivo or "FUENTES" in nombre_archivo:
            es_factura_usd = True
            moneda_inicial = "USD"  # Marcar inicialmente como USD
        
        # Leer el texto para buscar patrones similares
        reader = pypdf.PdfReader(archivo_pdf)
        texto = ""
        for pagina in reader.pages:
            texto += pagina.extract_text()
        
        # Verificación explícita para facturas en USD basada en el texto
        if "emitida en USD" in texto or "TOTAL USD:" in texto or (
            "cancelada en dicha moneda" in texto and "USD" in texto):
            es_factura_usd = True
            moneda_inicial = "USD"
        
        # Identificar tipo de factura
        tipo_factura = identificar_tipo_factura(texto)
        
        # Determinar moneda inicial si no se hizo por verificaciones anteriores
        if not es_factura_argentina and not es_factura_usd:
            if tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                moneda_inicial = "ARS"  # Forzar ARS para facturas A, B o electrónicas AFIP
            else:
                # Intentar detectar moneda normalmente
                moneda_inicial = detectar_moneda(texto)
        
        # 0. Buscar si hay un patrón similar ya aprendido (solo si no se ignoran patrones)
        if USE_PATTERN_MATCHING and not ignore_patterns:
            patron_similar = patrones_manager.encontrar_patron_similar(texto)
            if patron_similar:
                st.info(f"Usando patrón aprendido previamente")
                estructura = patron_similar['estructura']
                estructura['Nombre_Archivo'] = archivo_pdf.name
                
                # Si el patrón no tiene moneda, intentar asignar una
                if 'Moneda' not in estructura:
                    estructura['Moneda'] = moneda_inicial
                
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    estructura['Moneda'] = 'USD'
                elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                    estructura['Moneda'] = 'ARS'
                    
                estructura['Metodo'] = f"Patrón-{patron_similar['metodo_extraccion']}"
                return estructura
        
        # 1. Intentar con Tesseract OCR (alternativa gratuita a OpenAI)
        if USE_TESSERACT_OCR:
            archivo_pdf.seek(0)  # Reiniciar el puntero del archivo
            datos = extraer_con_tesseract_ocr(archivo_pdf)
            if datos and datos['Total'] > 0:
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    datos['Moneda'] = 'USD'
                elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                    datos['Moneda'] = 'ARS'
                else:
                    # Si no hay detección específica, usar la moneda inicial
                    datos['Moneda'] = moneda_inicial
                
                # Guardar el patrón aprendido
                patrones_manager.agregar_patron(datos, texto, 'TesseractOCR')
                return datos
        
        # 2. Intentar con Expresiones Regulares mejoradas
        archivo_pdf.seek(0)  # Reiniciar el puntero del archivo
        datos = extraer_con_regex(archivo_pdf)
        if datos and datos['Total'] > 0:
            # Forzar moneda según detecciones específicas
            if es_factura_usd:
                datos['Moneda'] = 'USD'
            elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
                datos['Moneda'] = 'ARS'
            else:
                # Si no hay detección específica, usar la moneda inicial
                datos['Moneda'] = moneda_inicial
            
            # Guardar el patrón aprendido
            patrones_manager.agregar_patron(datos, texto, 'RegEx')
            return datos
        
        # 3. Intentar específicamente con el extractor de facturas de viajes si es ese tipo
        if tipo_factura == "VIAJES":
            datos_viajes = extraer_datos_factura_viajes(texto)
            if datos_viajes and datos_viajes['Total'] > 0:
                datos_viajes['Nombre_Archivo'] = archivo_pdf.name
                
                # Forzar moneda según detecciones específicas
                if es_factura_usd:
                    datos_viajes['Moneda'] = 'USD'
                elif es_factura_argentina:
                    datos_viajes['Moneda'] = 'ARS'
                else:
                    datos_viajes['Moneda'] = moneda_inicial
                
                datos_viajes['Metodo'] = 'Especializado-Viajes'
                patrones_manager.agregar_patron(datos_viajes, texto, 'Especializado-Viajes')
                return datos_viajes
        
        # 4. Si todo falla, devolver datos vacíos
        resultado_fallido = {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': None,
            'Fecha': None,
            'Total': 0.0,
            'Gravado': 0.0,
            'IVA': 0.0,
            'No_Gravado': 0.0,
            'Exento': 0.0,
            'Moneda': moneda_inicial,  # Al menos tenemos una moneda detectada
            'Metodo': 'Fallido'
        }
        
        # Forzar moneda según detecciones específicas incluso si falló
        if es_factura_usd:
            resultado_fallido['Moneda'] = 'USD'
        elif es_factura_argentina or tipo_factura in ["TIPO_A", "TIPO_B", "ELECTRONICA_AFIP"]:
            resultado_fallido['Moneda'] = 'ARS'
            
        # Verificación final para "Bienes y srvs. no computables"
        if "Bienes y srvs. no computables" in texto and resultado_fallido['Total'] > 0:
            no_comp_match = re.search(r'(?:Bienes y srvs\. no computables|no computables para la det\. del Iva)(?:[^\d]*)\$?\s*([\d.,]+)', texto)
            if no_comp_match:
                try:
                    valor_str = no_comp_match.group(1).replace(".", "").replace(",", ".")
                    resultado_fallido['No_Gravado'] = float(valor_str)
                except:
                    pass
        
        return resultado_fallido
        
    except Exception as e:
        st.error(f"Error al procesar {archivo_pdf.name}: {str(e)}")
        return {
            'Nombre_Archivo': archivo_pdf.name,
            'Numero_Factura': None,
            'Fecha': None,
            'Total': 0.0,
            'Gravado': 0.0,
            'IVA': 0.0,
            'No_Gravado': 0.0,
            'Exento': 0.0,
            'Moneda': 'Desconocida',
            'Metodo': 'Error'
        }      
        

def aplicar_estilo_encabezado(ws, row, col_start, col_end):
    """Aplica estilos a las celdas de encabezado"""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def aplicar_estilo_datos(ws, row, col_start, col_end, es_moneda=False):
    """Aplica estilos a las celdas de datos"""
    data_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    data_alignment = Alignment(horizontal="right" if es_moneda else "left", vertical="center")
    
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = data_fill
        cell.alignment = data_alignment
        
        # Formato de moneda para columnas numéricas
        if es_moneda and col > 1:  # Excepto la columna de fecha
            cell.number_format = '"$"#,##0.00'


def ajustar_ancho_columnas(ws):
    """Ajusta el ancho de las columnas según el contenido"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def generar_excel(datos_list):
    """
    Genera un solo archivo Excel en memoria con una pestaña por factura y pestañas de resumen por moneda.
    """
    output = BytesIO()
    wb = Workbook()
    
    # Eliminar la hoja por defecto al principio
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Agrupar facturas por moneda
    facturas_por_moneda = {}
    for datos in datos_list:
        moneda = datos.get('Moneda', 'Desconocida')
        if moneda not in facturas_por_moneda:
            facturas_por_moneda[moneda] = []
        facturas_por_moneda[moneda].append(datos)
    
    # Crear una hoja de resumen global primero
    resumen_global = wb.create_sheet(title="Resumen Global")
    
    # Configurar encabezados del resumen global
    encabezados_resumen = ['FACTURA', 'FECHA', 'MONEDA', 'NO GRAVADO', 'EXENTO', 'GRAVADO', 'IVA', 'TOTAL', 'TOTAL A FACTURAR']
    for col_idx, header in enumerate(encabezados_resumen, 1):
        resumen_global.cell(row=1, column=col_idx, value=header)
    
    aplicar_estilo_encabezado(resumen_global, 1, 1, len(encabezados_resumen))
    
    # Añadir todas las facturas al resumen global
    fila_global = 2
    for moneda, facturas in facturas_por_moneda.items():
        for datos in facturas:
            resumen_global.cell(row=fila_global, column=1, value=datos['Numero_Factura'] or os.path.basename(datos['Nombre_Archivo']))
            resumen_global.cell(row=fila_global, column=2, value=datos['Fecha'])
            resumen_global.cell(row=fila_global, column=3, value=moneda)
            resumen_global.cell(row=fila_global, column=4, value=float(datos['No_Gravado']))
            resumen_global.cell(row=fila_global, column=5, value=float(datos['Exento']))
            resumen_global.cell(row=fila_global, column=6, value=float(datos['Gravado']))
            resumen_global.cell(row=fila_global, column=7, value=float(datos['IVA']))
            resumen_global.cell(row=fila_global, column=8, value=float(datos['Total']))
            resumen_global.cell(row=fila_global, column=9, value="")  # Celda en blanco para Total a Facturar
            
            # Aplicar formato monetario a las celdas numéricas
            for col in range(4, 9):
                resumen_global.cell(row=fila_global, column=col).number_format = '"$"#,##0.00'
            
            # También aplicar formato monetario a la celda de Total a Facturar
            resumen_global.cell(row=fila_global, column=9).number_format = '"$"#,##0.00'
            
            # Aplicar estilos
            aplicar_estilo_datos(resumen_global, fila_global, 1, 3, False)
            aplicar_estilo_datos(resumen_global, fila_global, 4, 9, True)  # Incluir la nueva columna
            
            fila_global += 1
    
    # Procesar cada moneda por separado
    for moneda, facturas in facturas_por_moneda.items():
        # Crear hoja de resumen para esta moneda
        nombre_resumen = f"Resumen {moneda}"
        resumen_sheet = wb.create_sheet(title=nombre_resumen)
        
        # Configurar encabezados del resumen
        encabezados_resumen = ['FACTURA', 'FECHA', 'NO GRAVADO', 'EXENTO', 'GRAVADO', 'IVA', 'TOTAL', 'TOTAL A FACTURAR']
        for col_idx, header in enumerate(encabezados_resumen, 1):
            resumen_sheet.cell(row=1, column=col_idx, value=header)
        
        aplicar_estilo_encabezado(resumen_sheet, 1, 1, len(encabezados_resumen))
        
        # Inicializar sumatorias para esta moneda
        total_no_gravado = 0.0
        total_exento = 0.0
        total_gravado = 0.0
        total_iva = 0.0
        total_total = 0.0
        
        # Procesar cada factura de esta moneda
        for i, datos in enumerate(facturas):
            # Crear hoja para la factura actual
            sheet_name = datos['Numero_Factura'] or os.path.basename(datos['Nombre_Archivo']).replace('.pdf', '')[:31]
            # Añadir un sufijo a la hoja para evitar nombres duplicados
            if moneda != 'Desconocida':
                sheet_name = f"{sheet_name}_{moneda}"
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Configurar encabezados de la factura
            headers = ['CONCEPTO', 'VALOR']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            aplicar_estilo_encabezado(ws, 1, 1, 2)
            
            # Agregar datos de factura en formato de tabla vertical
            conceptos = [
                'Número de Factura', 
                'Fecha', 
                'Moneda',
                'No Gravado', 
                'Exento', 
                'Gravado', 
                'IVA', 
                'TOTAL',
                'Total a facturar'  # Cambio de "Método de extracción" a "Total a facturar"
            ]
            
            # Asegurarse de que todos los valores numéricos sean float
            no_gravado = float(datos['No_Gravado'])
            exento = float(datos['Exento'])
            gravado = float(datos['Gravado'])
            iva = float(datos['IVA'])
            total = float(datos['Total'])
            
            valores = [
                datos['Numero_Factura'] or "N/D",
                datos['Fecha'] or "N/D",
                moneda,
                no_gravado,
                exento,
                gravado,
                iva,
                total,
                ""  # Celda en blanco para Total a facturar
            ]
            
            for j, (concepto, valor) in enumerate(zip(conceptos, valores), 2):
                ws.cell(row=j, column=1, value=concepto)
                ws.cell(row=j, column=2, value=valor)
                
                # Aplicar formato condicional
                if j >= 5 and j <= 8:  # Campos monetarios (No Gravado, Exento, Gravado, IVA, TOTAL)
                    ws.cell(row=j, column=2).number_format = '"$"#,##0.00'
                
                # También aplicar formato monetario a Total a facturar (celda 9)
                if j == 9:  # Total a facturar
                    ws.cell(row=j, column=2).number_format = '"$"#,##0.00'
            
            # Aplicar estilos a las celdas de datos
            for j in range(2, len(conceptos) + 2):
                aplicar_estilo_datos(ws, j, 1, 1, False)
                aplicar_estilo_datos(ws, j, 2, 2, j >= 5)  # Aplicar estilo monetario a todos los valores numéricos
            
            # Ajustar ancho de columnas
            ajustar_ancho_columnas(ws)
            
            # Agregar datos a la hoja de resumen de esta moneda
            resumen_sheet.cell(row=i+2, column=1, value=datos['Numero_Factura'] or os.path.basename(datos['Nombre_Archivo']))
            resumen_sheet.cell(row=i+2, column=2, value=datos['Fecha'])
            resumen_sheet.cell(row=i+2, column=3, value=no_gravado)
            resumen_sheet.cell(row=i+2, column=4, value=exento)
            resumen_sheet.cell(row=i+2, column=5, value=gravado)
            resumen_sheet.cell(row=i+2, column=6, value=iva)
            resumen_sheet.cell(row=i+2, column=7, value=total)
            resumen_sheet.cell(row=i+2, column=8, value="")  # Celda en blanco para Total a facturar
            
            # Aplicar formato monetario a las celdas de la hoja de resumen
            for col in range(3, 8):
                resumen_sheet.cell(row=i+2, column=col).number_format = '"$"#,##0.00'
            
            # También aplicar formato monetario a Total a facturar
            resumen_sheet.cell(row=i+2, column=8).number_format = '"$"#,##0.00'
            
            # Aplicar estilos a la fila de resumen
            aplicar_estilo_datos(resumen_sheet, i+2, 1, 2, False)
            aplicar_estilo_datos(resumen_sheet, i+2, 3, 8, True)
            
            # Acumular sumatorias con conversión explícita a float
            total_no_gravado += no_gravado
            total_exento += exento
            total_gravado += gravado
            total_iva += iva
            total_total += total
        
        # Agregar fila de totales en la hoja de resumen con valores calculados directamente
        ultima_fila = len(facturas) + 2
        resumen_sheet.cell(row=ultima_fila, column=1, value=f"TOTALES {moneda}")
        resumen_sheet.cell(row=ultima_fila, column=1).font = Font(bold=True)
        
        # Insertar totales calculados
        totales = [total_no_gravado, total_exento, total_gravado, total_iva, total_total]
        for idx, total in enumerate(totales):
            col = idx + 3  # Comenzando desde la columna 3 (NO GRAVADO)
            # Usar un valor numérico explícito
            resumen_sheet.cell(row=ultima_fila, column=col, value=float(total))
            resumen_sheet.cell(row=ultima_fila, column=col).font = Font(bold=True)
            resumen_sheet.cell(row=ultima_fila, column=col).number_format = '"$"#,##0.00'
            
            # Aplica un fondo diferente para los totales
            resumen_sheet.cell(row=ultima_fila, column=col).fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        
        # También aplicar estilo a las celdas adicionales de la fila de totales
        resumen_sheet.cell(row=ultima_fila, column=1).fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        resumen_sheet.cell(row=ultima_fila, column=2).fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        resumen_sheet.cell(row=ultima_fila, column=8).fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        
        # Ajustar ancho de columnas en la hoja de resumen
        ajustar_ancho_columnas(resumen_sheet)
    
    # Ajustar ancho de columnas en la hoja de resumen global
    ajustar_ancho_columnas(resumen_global)
    
    wb.save(output)
    output.seek(0)
    return output

def eliminar_duplicados_simple(datos_list):
    """Elimina duplicados comparando números de factura"""
    # Paso 1: Crear un diccionario de facturas por número
    facturas_por_numero = {}
    for datos in datos_list:
        numero = datos.get('Numero_Factura')
        if numero:
            if numero not in facturas_por_numero:
                facturas_por_numero[numero] = datos
    
    # Paso 2: Manejar facturas sin número
    facturas_sin_numero = [d for d in datos_list if not d.get('Numero_Factura')]
    
    # Paso 3: Combinar ambas listas
    return list(facturas_por_numero.values()) + facturas_sin_numero


def eliminar_duplicados_simple(datos_list):
    """Elimina duplicados comparando números de factura"""
    # Paso 1: Crear un diccionario de facturas por número
    facturas_por_numero = {}
    for datos in datos_list:
        numero = datos.get('Numero_Factura')
        if numero:
            if numero not in facturas_por_numero:
                facturas_por_numero[numero] = datos
    
    # Paso 2: Manejar facturas sin número
    facturas_sin_numero = [d for d in datos_list if not d.get('Numero_Factura')]
    
    # Paso 3: Combinar ambas listas
    return list(facturas_por_numero.values()) + facturas_sin_numero


def eliminar_duplicados_simple(datos_list):
    """Elimina duplicados comparando números de factura"""
    # Paso 1: Crear un diccionario de facturas por número
    facturas_por_numero = {}
    for datos in datos_list:
        numero = datos.get('Numero_Factura')
        if numero:
            if numero not in facturas_por_numero:
                facturas_por_numero[numero] = datos
    
    # Paso 2: Manejar facturas sin número
    facturas_sin_numero = [d for d in datos_list if not d.get('Numero_Factura')]
    
    # Paso 3: Combinar ambas listas
    return list(facturas_por_numero.values()) + facturas_sin_numero



def main():
    st.set_page_config(page_title="Extractor de Facturas", page_icon="📊", layout="wide")
    
    st.title("📊 Extractor Inteligente de Datos de Facturas")
    st.markdown("Sistema de extracción híbrido para procesar facturas y generar Excel con separación por monedas")
    
    # Inicializar el administrador de patrones
    patrones_manager = PatronesFacturas()
    
    # Sidebar con configuraciones
    with st.sidebar:
        st.header("Configuración")
        
        # Opciones de extracción
        st.subheader("Métodos de extracción")
        use_tesseract = st.checkbox("Usar OCR avanzado (Tesseract)", value=True)
        use_patterns = st.checkbox("Usar reconocimiento de patrones", value=True)
        
        # NUEVA OPCIÓN: Ignorar patrones guardados
        ignore_patterns = st.checkbox("Ignorar patrones guardados", value=False, 
                                    help="Activa esta opción para procesar las facturas sin usar patrones guardados")
        
        # Para activar/desactivar características
        global USE_TESSERACT_OCR, USE_PATTERN_MATCHING
        USE_TESSERACT_OCR = use_tesseract
        USE_PATTERN_MATCHING = use_patterns
        
        # Modo depuración
        st.subheader("Depuración")
        debug_mode = st.checkbox("Modo depuración", value=False)
        
        # Información sobre patrones aprendidos
        st.subheader("Base de conocimiento")
        patrones_count = len(patrones_manager.patrones)
        st.write(f"Patrones aprendidos: {patrones_count}")
        
        if st.button("Reiniciar base de conocimiento"):
            patrones_manager.patrones = {}
            patrones_manager.guardar_patrones()
            st.success("Base de conocimiento reiniciada")
    
    # Contenido principal
    with st.expander("ℹ️ Instrucciones de uso", expanded=False):
        st.markdown("""
        1. Selecciona uno o varios archivos PDF de facturas usando el selector de archivos.
        2. Haz clic en 'Procesar Facturas' para extraer los datos.
        3. El sistema utilizará una combinación de tecnologías:
           - OCR avanzado con Tesseract (gratuito)
           - Aprendizaje automático de patrones
           - Expresiones regulares como respaldo
           - **NUEVO**: Reconocimiento especializado para facturas de agencias de viajes
           - **NUEVO**: Detección automática de moneda mejorada para facturas argentinas
        4. Detecta automáticamente la moneda de cada factura
        5. Revisa la vista previa de los datos extraídos.
        6. Puedes editar manualmente cualquier valor incorrecto.
        7. Descarga el Excel generado con pestañas separadas por moneda.
        """)
    
    # Verificar si Tesseract está instalado
    if USE_TESSERACT_OCR:
        try:
            import pytesseract
            pytesseract.get_tesseract_version()
        except Exception as e:
            st.warning(f"""
            ⚠️ Tesseract OCR no está instalado o configurado correctamente: {str(e)}
            
            Para usarlo necesitas:
            1. Instalar Tesseract OCR en tu sistema
            2. Instalar los paquetes Python: `pip install pytesseract pdf2image opencv-python`
            
            El programa funcionará con expresiones regulares como respaldo.
            """)
    
    # Subir archivos
    archivos_pdf = st.file_uploader("Selecciona archivos PDF de facturas", type="pdf", accept_multiple_files=True)
    
    if archivos_pdf:
        st.info(f"Se han cargado {len(archivos_pdf)} archivo(s). Presiona 'Procesar Facturas' para continuar.")
        
        if st.button("Procesar Facturas", type="primary"):
            datos_list = []
            
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            for idx, archivo in enumerate(archivos_pdf):
                progress_text.text(f"Procesando {archivo.name}... ({idx+1}/{len(archivos_pdf)})")
                progress_bar.progress((idx + 0.5) / len(archivos_pdf))
                
                # Aplicar el sistema de cascada de extracción
                with st.spinner(f"Analizando {archivo.name}..."):
                    # Usar el parámetro ignore_patterns
                    datos = extraer_datos_pdf(archivo, patrones_manager, ignore_patterns=ignore_patterns)
                    datos_list.append(datos)
                
                progress_bar.progress((idx + 1) / len(archivos_pdf))
            
            progress_text.text("¡Procesamiento completado!")
            
            # NUEVA LÍNEA: Eliminar duplicados antes de mostrar resultados
            datos_list = eliminar_duplicados_simple(datos_list)
            
            # Mostrar vista previa de datos extraídos
            if datos_list:
                st.success(f"Se procesaron {len(datos_list)} facturas correctamente.")
                
                # Agrupar facturas por moneda para mostrar en pestañas
                facturas_por_moneda = {}
                for datos in datos_list:
                    moneda = datos.get('Moneda', 'Desconocida')
                    if moneda not in facturas_por_moneda:
                        facturas_por_moneda[moneda] = []
                    facturas_por_moneda[moneda].append(datos)
                
                # Crear pestañas para cada moneda
                tab_global, *tabs_moneda = st.tabs(["Todas las facturas"] + [f"Facturas en {m}" for m in facturas_por_moneda.keys()])
                
                with tab_global:
                    # Crear DataFrame para edición global
                    df_editable = pd.DataFrame([
                        {
                            'Factura': d['Numero_Factura'] or os.path.basename(d['Nombre_Archivo']),
                            'Fecha': d['Fecha'] or "",
                            'Moneda': d.get('Moneda', 'Desconocida'),
                            'No Gravado': float(d['No_Gravado']),
                            'Exento': float(d['Exento']),
                            'Gravado': float(d['Gravado']),
                            'IVA': float(d['IVA']),
                            'Total': float(d['Total']),
                            
                        } for d in datos_list
                    ])
                    
                    # Permitir edición de los valores extraídos
                    st.subheader("Editar valores extraídos")
                    st.write("Puedes editar cualquier valor incorrecto antes de generar el Excel.")
                    
                    # Crear una copia editable del DataFrame
                    edited_df = st.data_editor(
                        df_editable,
                        use_container_width=True,
                        column_config={
                            "Moneda": st.column_config.SelectboxColumn(
                                "Moneda",
                                options=["ARS", "USD", "EUR", "Desconocida"],
                                required=True
                            ),
                            "No Gravado": st.column_config.NumberColumn(
                                "No Gravado",
                                format="%.2f",
                                step=0.01,
                            ),
                            "Exento": st.column_config.NumberColumn(
                                "Exento",
                                format="%.2f",
                                step=0.01,
                            ),
                            "Gravado": st.column_config.NumberColumn(
                                "Gravado",
                                format="%.2f",
                                step=0.01,
                            ),
                            "IVA": st.column_config.NumberColumn(
                                "IVA",
                                format="%.2f",
                                step=0.01,
                            ),
                            "Total": st.column_config.NumberColumn(
                                "Total",
                                format="%.2f",
                                step=0.01,
                            )
                        }
                    )
                    
                    # Actualizar los datos con los valores editados
                    for i, (_, row) in enumerate(edited_df.iterrows()):
                        if i < len(datos_list):  # Verificar que no excedemos el índice
                            datos_list[i]['No_Gravado'] = row['No Gravado']
                            datos_list[i]['Exento'] = row['Exento']
                            datos_list[i]['Gravado'] = row['Gravado']
                            datos_list[i]['IVA'] = row['IVA']
                            datos_list[i]['Total'] = row['Total']
                            datos_list[i]['Fecha'] = row['Fecha']
                            datos_list[i]['Moneda'] = row['Moneda']
                    
                    # Recalcular las agrupaciones por moneda después de editar
                    facturas_por_moneda = {}
                    for datos in datos_list:
                        moneda = datos.get('Moneda', 'Desconocida')
                        if moneda not in facturas_por_moneda:
                            facturas_por_moneda[moneda] = []
                        facturas_por_moneda[moneda].append(datos)
                
                # Mostrar resumen de totales por moneda en cada pestaña
                for i, (moneda, facturas) in enumerate(facturas_por_moneda.items()):
                    if i < len(tabs_moneda):  # Asegurarse de que no excedemos el número de pestañas
                        with tabs_moneda[i]:
                            # Crear DataFrame para esta moneda
                            df_moneda = pd.DataFrame([
                                {
                                    'Factura': d['Numero_Factura'] or os.path.basename(d['Nombre_Archivo']),
                                    'Fecha': d['Fecha'] or "",
                                    'No Gravado': float(d['No_Gravado']),
                                    'Exento': float(d['Exento']),
                                    'Gravado': float(d['Gravado']),
                                    'IVA': float(d['IVA']),
                                    'Total': float(d['Total'])
                                } for d in facturas
                            ])
                            
                            st.subheader(f"Facturas en {moneda}")
                            st.dataframe(df_moneda, use_container_width=True)
                            
                            # Calcular totales para esta moneda
                            totales_moneda = {
                                'No Gravado': df_moneda['No Gravado'].sum(),
                                'Exento': df_moneda['Exento'].sum(),
                                'Gravado': df_moneda['Gravado'].sum(),
                                'IVA': df_moneda['IVA'].sum(),
                                'Total': df_moneda['Total'].sum()
                            }
                            
                            # Mostrar resumen de totales para esta moneda
                            st.subheader(f"Resumen totales en {moneda}")
                            col1, col2, col3, col4, col5 = st.columns(5)
                            col1.metric("No Gravado", f"${totales_moneda['No Gravado']:.2f}")
                            col2.metric("Exento", f"${totales_moneda['Exento']:.2f}")
                            col3.metric("Gravado", f"${totales_moneda['Gravado']:.2f}")
                            col4.metric("IVA", f"${totales_moneda['IVA']:.2f}")
                            col5.metric("TOTAL", f"${totales_moneda['Total']:.2f}")
                
                # Generar y descargar Excel
                excel_file = generar_excel(datos_list)
                
                # Botón para descargar el Excel
                st.download_button(
                    label="Descargar Excel",
                    data=excel_file,
                    file_name=f"Facturas_por_Moneda_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.error("No se pudo extraer datos de ninguna factura.")

if __name__ == "__main__":
    main()
